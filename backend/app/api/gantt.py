"""
Gantt / WBS API — Work Breakdown Structure scheduling and progress tracking.

Hierarchy: Category -> Activity -> Sub-Activity (leaf, stored as GanttTask).
Activity- and Category-level rows are never stored; they are computed live
by rolling up their child Sub-Activities:
    planned_start = MIN(sub-activities' planned_start)
    planned_end   = MAX(sub-activities' planned_end)
    planned_pct   = MEAN(sub-activities' planned_pct)   -- equal weight
    actual_pct    = MEAN(sub-activities' actual_pct)    -- equal weight
This is proven to match a native Excel PivotTable's Category-level subtotal
exactly (verified against the reference workbook) when the aggregation is
applied directly to the leaf rows rather than nested through intermediate
averages.

Per Sub-Activity:
    planned_pct = NETWORKDAYS-based linear ramp from planned_start to
                  planned_end, evaluated against today (0 before start,
                  1 after end) -- mirrors the reference workbook's formula.
    actual_pct  = looked up from `status` via STATUS_PERCENT_MAP -- the
                  user picks a status, never types a percentage directly.

Endpoints:
  GET    /api/gantt/statuses                    -> fixed WBS status list + percentages
  GET    /api/gantt/{project_id}/tasks          -> flat list of sub-activities
  POST   /api/gantt/{project_id}/tasks          -> add a sub-activity
  PATCH  /api/gantt/tasks/{task_id}             -> edit a sub-activity
  DELETE /api/gantt/tasks/{task_id}             -> remove a sub-activity
  POST   /api/gantt/tasks/{task_id}/progress    -> log a status update
  GET    /api/gantt/tasks/{task_id}/logs        -> progress history for one sub-activity
  GET    /api/gantt/tasks/{task_id}/revisions   -> schedule/WBS change history
  GET    /api/gantt/{project_id}/rollup         -> Category -> Activity rollup (dashboard)
  GET    /api/gantt/{project_id}/s-curve        -> planned vs actual over time (project-wide)
  GET    /api/gantt/{project_id}/progress-summary -> single {planned_progress, actual_progress}
"""

from datetime import date, datetime, timedelta
from typing import Optional, List
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.db.database import get_db
from app.models import GanttTask, GanttProgressLog, Project, Resource, Milestone, PendingDateChange
from app.auth import get_current_resource, require_schedule_edit_access, can_view_schedule, can_edit_schedule, is_management

gantt_router = APIRouter(prefix="/api/gantt", tags=["Gantt"])


# ── Fixed WBS status list (source of truth — matches the WBS reference sheet) ─
# Order matters for the dropdown; percent is what "Actual Progress" resolves to.
STATUS_PERCENT_MAP = [
    {"status": "Not Started",          "percent": 0},
    {"status": "Assigned",             "percent": 0},
    {"status": "In Progress",          "percent": 30},
    {"status": "Testing",              "percent": 60},
    {"status": "Pending Review",       "percent": 80},
    {"status": "Pending Sign-Off",     "percent": 80},
    {"status": "Pending Deployment",   "percent": 90},
    {"status": "Deferred",             "percent": 100},
    {"status": "Completed / Sign-Off", "percent": 100},
]
_STATUS_LOOKUP = {row["status"]: row["percent"] for row in STATUS_PERCENT_MAP}
_VALID_STATUSES = set(_STATUS_LOOKUP.keys())


def _percent_for_status(status: Optional[str]) -> float:
    if not status:
        return 0.0
    return float(_STATUS_LOOKUP.get(status, 0))


def _networkdays(start: datetime, end: datetime) -> int:
    """Count working days (Mon-Fri) between two dates, inclusive of both ends."""
    if end < start:
        return 0
    days = 0
    cursor = start
    while cursor <= end:
        if cursor.weekday() < 5:  # 0=Mon .. 4=Fri
            days += 1
        cursor += timedelta(days=1)
    return days


def _planned_pct_for(planned_start: Optional[str], planned_end: Optional[str], as_of: Optional[datetime] = None) -> float:
    """
    NETWORKDAYS-based linear ramp, mirroring the reference workbook:
      0%   if as_of < start
      100% if as_of >= end
      else (working days elapsed) / (total working days in range)
    """
    if not planned_start or not planned_end:
        return 0.0
    s = datetime.strptime(planned_start[:10], "%Y-%m-%d")
    e = datetime.strptime(planned_end[:10], "%Y-%m-%d")
    today = as_of or datetime.today()
    if today < s:
        return 0.0
    if today >= e or e <= s:
        return 100.0
    elapsed = _networkdays(s, today)
    total = _networkdays(s, e)
    if total <= 0:
        return 100.0
    return round(min(100.0, max(0.0, (elapsed / total) * 100)), 1)


# ── Pydantic models ───────────────────────────────────────────────────────────

class GanttTaskIn(BaseModel):
    category:      str
    activity:      str
    task_name:     str    # Sub-Activity name
    assigned_to:   Optional[str] = None
    planned_start: Optional[str] = None
    planned_end:   Optional[str] = None
    status:        str = "Not Started"


class GanttTaskUpdate(BaseModel):
    category:      Optional[str] = None
    activity:      Optional[str] = None
    task_name:     Optional[str] = None
    assigned_to:   Optional[str] = None
    planned_start: Optional[str] = None
    planned_end:   Optional[str] = None


class ProgressLogIn(BaseModel):
    description: Optional[str] = None
    status:      str


# ── Serializers ────────────────────────────────────────────────────────────────

def pending_date_change_to_dict(p: PendingDateChange) -> dict:
    return {
        "id": p.id, "gantt_task_id": p.gantt_task_id, "project_id": p.project_id,
        "requested_planned_start": p.requested_planned_start,
        "requested_planned_end": p.requested_planned_end,
        "previous_planned_start": p.previous_planned_start,
        "previous_planned_end": p.previous_planned_end,
        "requested_by_name": p.requested_by_name,
        "requested_at": p.requested_at.isoformat() if p.requested_at else None,
        "status": p.status,
        "review_note": p.review_note,
    }


def task_to_dict(t: GanttTask, pending_map: Optional[dict] = None) -> dict:
    return {
        "id":               t.entity_id,   # stable across versions — never the per-version surrogate t.id
        "project_id":       t.project_id,
        "category":         t.category,
        "activity":         t.activity,
        "task_name":        t.task_name,
        "assigned_to":      t.assigned_to,
        "planned_start":    t.planned_start,
        "planned_end":      t.planned_end,
        "status":           t.status,
        "percent_complete": t.percent_complete,           # actual %, from status
        "planned_percent":  _planned_pct_for(t.planned_start, t.planned_end),
        "sort_order":       t.sort_order,
        "version":          t.version,
        "effective_date":   t.effective_date.isoformat() if t.effective_date else None,
        "pending_date_change": (pending_map or {}).get(t.entity_id),
    }


def log_to_dict(l: GanttProgressLog) -> dict:
    return {
        "id":               l.id,
        "gantt_task_id":    l.gantt_task_id,
        "log_date":         l.log_date,
        "description":      l.description,
        "status":           l.status,
        "percent_complete": l.percent_complete,
    }


# ── GET fixed status list ──────────────────────────────────────────────────────

@gantt_router.get("/statuses")
def get_statuses():
    """The fixed WBS status list, in display order, with each status's resolved %."""
    return STATUS_PERCENT_MAP


# ── GET tasks (flat list of sub-activities) ───────────────────────────────────

def _project_for_view_or_403(project_id: int, db: Session, current_resource: Optional[Resource]) -> Project:
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if not can_view_schedule(current_resource, project, db):
        raise HTTPException(status_code=403, detail="You don't have access to this project's schedule.")
    return project


def _get_current_task(db: Session, entity_id: int) -> Optional[GanttTask]:
    return db.query(GanttTask).filter(GanttTask.entity_id == entity_id, GanttTask.is_current == True).first()


def _new_task_version(db: Session, current: GanttTask, **changes) -> GanttTask:
    """
    SCD Type 2 write: close out `current` (is_current=False, end_date=now)
    and insert a brand-new row — same entity_id, version+1 — carrying every
    existing field forward except whatever's passed in `changes`. The old
    row is never mutated or deleted; it simply becomes history. Returns the
    new current row (not yet committed — caller commits).
    """
    now = datetime.utcnow()
    current.is_current = False
    current.end_date = now

    new = GanttTask(
        entity_id=current.entity_id,
        version=current.version + 1,
        effective_date=now,
        end_date=None,
        is_current=True,
        project_id=current.project_id,
        category=changes.get("category", current.category),
        activity=changes.get("activity", current.activity),
        task_name=changes.get("task_name", current.task_name),
        assigned_to=changes.get("assigned_to", current.assigned_to),
        planned_start=changes.get("planned_start", current.planned_start),
        planned_end=changes.get("planned_end", current.planned_end),
        status=changes.get("status", current.status),
        percent_complete=changes.get("percent_complete", current.percent_complete),
        sort_order=changes.get("sort_order", current.sort_order),
    )
    db.add(new)
    return new


@gantt_router.get("/{project_id}/tasks")
def list_tasks(
    project_id: int, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    project = _project_for_view_or_403(project_id, db, current_resource)
    tasks = (
        db.query(GanttTask)
        .filter(GanttTask.project_id == project_id, GanttTask.is_current == True)
        .order_by(GanttTask.sort_order, GanttTask.entity_id)
        .all()
    )
    pending_rows = db.query(PendingDateChange).filter(
        PendingDateChange.project_id == project_id, PendingDateChange.status == "pending"
    ).all()
    pending_map = {p.gantt_task_id: pending_date_change_to_dict(p) for p in pending_rows}
    return {"tasks": [task_to_dict(t, pending_map) for t in tasks], "wbs_uploaded": project.wbs_uploaded}


# ── POST task (add a sub-activity) ────────────────────────────────────────────

@gantt_router.post("/{project_id}/tasks", status_code=201)
def create_task(
    project_id: int, body: GanttTaskIn, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    require_schedule_edit_access(current_resource, project, db)
    if body.status not in _VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(_VALID_STATUSES)}")

    max_order = db.query(GanttTask).filter(GanttTask.project_id == project_id, GanttTask.is_current == True).count()
    t = GanttTask(
        project_id=project_id,
        category=body.category,
        activity=body.activity,
        task_name=body.task_name,
        assigned_to=body.assigned_to,
        planned_start=body.planned_start or None,
        planned_end=body.planned_end or None,
        status=body.status,
        percent_complete=_percent_for_status(body.status),
        sort_order=max_order,
        version=1, is_current=True,
    )
    db.add(t)
    db.flush()          # assigns t.id so it can become its own stable entity_id
    t.entity_id = t.id  # version 1 IS the entity — every later version reuses this id
    # Manually adding a task claims this project's WBS the same way an Excel
    # upload does — Upload WBS Excel closes off from here on, matching the
    # existing "only one entry method per project" intent that the upload
    # side already enforced but manual entry never did.
    project.wbs_uploaded = True
    db.commit()
    db.refresh(t)
    return task_to_dict(t)


# ── PATCH task (edit WBS placement / schedule — NOT status; use /progress for that) ─

@gantt_router.patch("/tasks/{task_id}")
def update_task(
    task_id: int, body: GanttTaskUpdate, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    """task_id here is the stable entity_id, not a specific version's surrogate id."""
    t = _get_current_task(db, task_id)
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")
    project = db.query(Project).filter(Project.id == t.project_id).first()
    if project:
        require_schedule_edit_access(current_resource, project, db)

    wants_date_change = (
        (body.planned_start is not None and body.planned_start != t.planned_start) or
        (body.planned_end is not None and body.planned_end != t.planned_end)
    )
    # Management edits dates immediately (they're the approver, not a requester).
    # Anyone else's date change on an EXISTING row goes to Management for
    # approval instead of applying — every other field still updates right away.
    gate_dates = wants_date_change and not is_management(current_resource)
    date_changing = wants_date_change and not gate_dates

    # Collect everything that should land on a NEW version right now. Gated
    # date changes are deliberately excluded here — the current version's
    # dates must stay untouched until Management approves the request below.
    changes = {}
    if body.category is not None and body.category != t.category:       changes["category"] = body.category
    if body.activity is not None and body.activity != t.activity:       changes["activity"] = body.activity
    if body.task_name is not None and body.task_name != t.task_name:    changes["task_name"] = body.task_name
    if body.assigned_to is not None and body.assigned_to != t.assigned_to: changes["assigned_to"] = body.assigned_to
    if date_changing:
        if body.planned_start is not None: changes["planned_start"] = body.planned_start or None
        if body.planned_end is not None:   changes["planned_end"] = body.planned_end or None

    if changes:
        t = _new_task_version(db, t, **changes)

    pending_result = None
    if gate_dates:
        # Replace any existing still-pending request on this task rather than queuing
        existing = db.query(PendingDateChange).filter(
            PendingDateChange.gantt_task_id == t.entity_id, PendingDateChange.status == "pending"
        ).first()
        req_start = body.planned_start if body.planned_start is not None else t.planned_start
        req_end   = body.planned_end if body.planned_end is not None else t.planned_end
        if existing:
            existing.requested_planned_start = req_start
            existing.requested_planned_end = req_end
            existing.previous_planned_start = t.planned_start
            existing.previous_planned_end = t.planned_end
            existing.requested_by = current_resource.id if current_resource else None
            existing.requested_by_name = current_resource.name if current_resource else None
            pdc = existing
        else:
            pdc = PendingDateChange(
                gantt_task_id=t.entity_id, project_id=t.project_id,
                requested_planned_start=req_start, requested_planned_end=req_end,
                previous_planned_start=t.planned_start, previous_planned_end=t.planned_end,
                requested_by=current_resource.id if current_resource else None,
                requested_by_name=current_resource.name if current_resource else None,
                status="pending",
            )
            db.add(pdc)
        db.flush()
        pending_result = pending_date_change_to_dict(pdc)

    db.commit()
    db.refresh(t)
    result = task_to_dict(t)
    result["pending_date_change"] = pending_result
    return result


# ── DELETE task ────────────────────────────────────────────────────────────────

@gantt_router.delete("/tasks/{task_id}", status_code=204)
def delete_task(
    task_id: int, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    """
    Soft delete: closes out the current version (is_current=False, end_date=now).
    No new version is inserted — there's no successor, the task is just gone.
    The full history up to the deletion point is still there if ever needed.
    """
    t = _get_current_task(db, task_id)
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")
    project = db.query(Project).filter(Project.id == t.project_id).first()
    if project:
        require_schedule_edit_access(current_resource, project, db)
    t.is_current = False
    t.end_date = datetime.utcnow()
    db.commit()
    return None


# ── POST progress log (choose a status — percent is resolved automatically) ──

@gantt_router.post("/tasks/{task_id}/progress", status_code=201)
def log_progress(
    task_id: int, body: ProgressLogIn, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    """task_id here is the stable entity_id, not a specific version's surrogate id."""
    t = _get_current_task(db, task_id)
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")
    project = db.query(Project).filter(Project.id == t.project_id).first()
    if project:
        require_schedule_edit_access(current_resource, project, db)
    if body.status not in _VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(_VALID_STATUSES)}")

    resolved_pct = _percent_for_status(body.status)
    log = GanttProgressLog(
        gantt_task_id=t.entity_id,
        log_date=date.today().isoformat(),
        description=body.description,
        status=body.status,
        percent_complete=resolved_pct,
    )
    db.add(log)
    # status/percent_complete are versioned fields — a progress update creates
    # a new version too, same as any other tracked field change.
    t = _new_task_version(db, t, status=body.status, percent_complete=resolved_pct)
    db.commit()
    db.refresh(log)
    db.refresh(t)
    return {"log": log_to_dict(log), "task": task_to_dict(t)}


# ── GET progress logs for a task ──────────────────────────────────────────────

@gantt_router.get("/tasks/{task_id}/logs")
def get_task_logs(task_id: int, db: Session = Depends(get_db)):
    logs = db.query(GanttProgressLog).filter(GanttProgressLog.gantt_task_id == task_id).order_by(GanttProgressLog.log_date).all()
    return [log_to_dict(l) for l in logs]


# ── GET schedule/WBS version history for a task ───────────────────────────────
# Replaces the old separate revisions table entirely — under SCD Type 2 the
# full history IS just every row sharing this entity_id, oldest first.

@gantt_router.get("/tasks/{task_id}/revisions")
def get_task_revisions(task_id: int, db: Session = Depends(get_db)):
    versions = (
        db.query(GanttTask)
        .filter(GanttTask.entity_id == task_id)
        .order_by(GanttTask.version.desc())
        .all()
    )
    return [
        {
            "id": v.id, "version": v.version,
            "category": v.category, "activity": v.activity, "task_name": v.task_name,
            "assigned_to": v.assigned_to, "planned_start": v.planned_start, "planned_end": v.planned_end,
            "status": v.status, "percent_complete": v.percent_complete,
            "is_current": v.is_current,
            "effective_date": v.effective_date.isoformat() if v.effective_date else None,
            "end_date": v.end_date.isoformat() if v.end_date else None,
        }
        for v in versions
    ]


# ── Rollup helpers ──────────────────────────────────────────────────────────────

def _rollup_group(tasks: List[GanttTask]) -> dict:
    """MIN/MAX dates + MEAN planned/actual % directly over the given leaf tasks."""
    if not tasks:
        return {"planned_start": None, "planned_end": None, "planned_pct": 0, "actual_pct": 0, "count": 0}
    starts = [t.planned_start for t in tasks if t.planned_start]
    ends   = [t.planned_end   for t in tasks if t.planned_end]
    planned_pcts = [_planned_pct_for(t.planned_start, t.planned_end) for t in tasks]
    actual_pcts  = [t.percent_complete or 0 for t in tasks]
    return {
        "planned_start": min(starts) if starts else None,
        "planned_end":   max(ends) if ends else None,
        "planned_pct":   round(sum(planned_pcts) / len(planned_pcts), 1) if planned_pcts else 0,
        "actual_pct":    round(sum(actual_pcts) / len(actual_pcts), 1) if actual_pcts else 0,
        "count":         len(tasks),
    }


# ── GET rollup (Category -> Activity, for the dashboard) ─────────────────────

@gantt_router.get("/{project_id}/rollup")
def get_rollup(
    project_id: int, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    """
    Category -> Activity rollup, each level computed directly from its leaf
    Sub-Activities (not chained through intermediate averages).
    """
    _project_for_view_or_403(project_id, db, current_resource)
    tasks = db.query(GanttTask).filter(GanttTask.project_id == project_id, GanttTask.is_current == True).all()
    if not tasks:
        return {"project": _rollup_group([]), "categories": []}

    categories: dict = {}
    for t in tasks:
        categories.setdefault(t.category or "Uncategorised", []).append(t)

    cat_rows = []
    for cat_name, cat_tasks in categories.items():
        activities: dict = {}
        for t in cat_tasks:
            activities.setdefault(t.activity or "Unassigned", []).append(t)
        activity_rows = [
            {"activity": act_name, **_rollup_group(act_tasks)}
            for act_name, act_tasks in activities.items()
        ]
        cat_rows.append({
            "category": cat_name,
            **_rollup_group(cat_tasks),
            "activities": activity_rows,
        })

    return {
        "project": _rollup_group(tasks),
        "categories": cat_rows,
    }


# ── GET S-curve (planned vs actual over time, project-wide) ──────────────────

@gantt_router.get("/{project_id}/s-curve")
def get_s_curve(
    project_id: int, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    _project_for_view_or_403(project_id, db, current_resource)
    tasks = db.query(GanttTask).filter(GanttTask.project_id == project_id, GanttTask.is_current == True).all()
    if not tasks:
        return {"months": [], "planned": [], "actual": []}

    starts = [datetime.strptime(t.planned_start[:10], "%Y-%m-%d") for t in tasks if t.planned_start]
    ends   = [datetime.strptime(t.planned_end[:10],   "%Y-%m-%d") for t in tasks if t.planned_end]
    if not starts or not ends:
        return {"months": [], "planned": [], "actual": []}

    range_start = min(starts)
    range_end   = max(max(ends), datetime.today())

    # Pre-fetch logs per task for actual-at-a-point-in-time lookups. Keyed by
    # entity_id (not the current version's surrogate id) — a progress log
    # belongs to the task as a concept, tracked the same way everywhere else.
    entity_ids = [t.entity_id for t in tasks]
    all_logs = db.query(GanttProgressLog).filter(GanttProgressLog.gantt_task_id.in_(entity_ids)).all() if entity_ids else []
    task_logs: dict = {}
    for l in all_logs:
        task_logs.setdefault(l.gantt_task_id, []).append((datetime.strptime(l.log_date, "%Y-%m-%d"), l.percent_complete))
    for k in task_logs:
        task_logs[k].sort(key=lambda x: x[0])

    def actual_pct_at(t: GanttTask, sample_date: datetime) -> float:
        logs = task_logs.get(t.entity_id, [])
        latest = 0.0
        for log_date, pct in logs:
            if log_date <= sample_date:
                latest = pct
            else:
                break
        return latest

    months: List[str] = []
    planned_series: List[float] = []
    actual_series: List[float] = []

    cursor = datetime(range_start.year, range_start.month, 1)
    end_marker = datetime(range_end.year, range_end.month, 1)

    while cursor <= end_marker:
        if cursor.month == 12:
            month_end = datetime(cursor.year, 12, 31)
        else:
            month_end = datetime(cursor.year, cursor.month + 1, 1) - timedelta(days=1)

        planned_vals = [_planned_pct_for(t.planned_start, t.planned_end, as_of=month_end) for t in tasks]
        actual_vals  = [actual_pct_at(t, month_end) for t in tasks]

        months.append(cursor.strftime("%b %Y"))
        planned_series.append(round(sum(planned_vals) / len(planned_vals), 1) if planned_vals else 0)
        actual_series.append(round(sum(actual_vals) / len(actual_vals), 1) if actual_vals else 0)

        if cursor.month == 12:
            cursor = datetime(cursor.year + 1, 1, 1)
        else:
            cursor = datetime(cursor.year, cursor.month + 1, 1)

    return {"months": months, "planned": planned_series, "actual": actual_series}


# ── GET overall progress summary (for Portfolio KPI / project header) ────────

@gantt_router.get("/{project_id}/progress-summary")
def get_progress_summary(
    project_id: int, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    _project_for_view_or_403(project_id, db, current_resource)
    tasks = db.query(GanttTask).filter(GanttTask.project_id == project_id, GanttTask.is_current == True).all()
    rollup = _rollup_group(tasks)
    return {"planned_progress": rollup["planned_pct"], "actual_progress": rollup["actual_pct"]}


# ── POST bulk upload from Excel (WBS sheet format) ─────────────────────────────

_EXPECTED_HEADERS = ["item", "category", "activity", "sub-activity"]


def _find_header_row(ws) -> Optional[int]:
    """
    Scan the first 40 rows for one containing Item/Category/Activity/Sub-Activity
    as adjacent column headers (robust to exact row position varying by template).
    """
    for r in range(1, min(41, ws.max_row + 1)):
        vals = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(1, 6)]
        if all(h in vals for h in _EXPECTED_HEADERS):
            return r
    return None


def _find_wbs_sheet(wb):
    """Prefer a sheet whose name contains 'WBS'; fall back to the active sheet."""
    for name in wb.sheetnames:
        if "wbs" in name.lower():
            return wb[name]
    return wb.active


def _cell_date_str(value) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    # Occasionally dates land as text (e.g. "2026-01-05") — pass through if it looks date-like
    s = str(value).strip()
    return s[:10] if len(s) >= 8 else None


@gantt_router.post("/{project_id}/upload")
async def upload_wbs_excel(
    project_id: int, file: UploadFile = File(...), db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    """
    Bulk-import Category / Activity / Sub-Activity / Planned Start / Planned End
    from an uploaded WBS Excel file (same layout as the reference WBS sheet:
    Item, Category, Activity, Sub-Activity, Start Date, Start Day, End Date, End Day —
    'Start Day'/'End Day' are informational weekday labels and are ignored here).

    Upsert behaviour: matches existing sub-activities by (category, activity,
    task_name). If found, only WBS placement + dates are updated (a revision
    snapshot is saved) — status/progress already logged is left untouched.
    If not found, a new sub-activity is created with status 'Not Started'.
    Rows already in the database but missing from the uploaded file are
    NOT deleted — re-uploading only adds/updates, never removes.
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    require_schedule_edit_access(current_resource, project, db)

    if project.wbs_uploaded:
        raise HTTPException(
            status_code=400,
            detail="A WBS Excel has already been uploaded for this project. "
                   "Further changes should be made directly in the Schedule tab."
        )

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx or .xlsm file")

    try:
        import openpyxl
        contents = await file.read()
        import io
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read the Excel file: {e}")

    ws = _find_wbs_sheet(wb)
    header_row = _find_header_row(ws)
    if header_row is None:
        raise HTTPException(
            status_code=400,
            detail="Could not find a header row with Item / Category / Activity / Sub-Activity columns. "
                   "Please keep the same column layout as the WBS reference sheet."
        )

    # Map header labels to column indices (case-insensitive, trimmed)
    col_map: dict = {}
    for c in range(1, ws.max_column + 1):
        label = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
        if label:
            col_map[label] = c

    required = ["category", "activity", "sub-activity"]
    missing = [h for h in required if h not in col_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required column(s): {', '.join(missing)}")

    col_category = col_map["category"]
    col_activity = col_map["activity"]
    col_subact   = col_map["sub-activity"]
    col_start    = col_map.get("start date") or col_map.get("planned start date")
    col_end      = col_map.get("end date") or col_map.get("planned end date")
    col_assigned = col_map.get("assigned to")

    # Existing tasks for this project, keyed for upsert matching.
    # Snapshot taken BEFORE this upload — matches are consumed one-at-a-time so
    # that a repeated (category, activity, sub-activity) triple within the same
    # file correctly becomes additional new rows instead of colliding with a
    # not-yet-flushed row created earlier in this same pass. Only matches
    # against CURRENT versions — a superseded historical row should never be
    # treated as "the" existing sub-activity to update.
    existing = db.query(GanttTask).filter(GanttTask.project_id == project_id, GanttTask.is_current == True).all()
    existing_by_key: dict = {}
    for t in existing:
        existing_by_key.setdefault((t.category, t.activity, t.task_name), []).append(t)
    max_sort = len(existing)

    created, updated, skipped = 0, 0, 0
    errors: List[str] = []

    for r in range(header_row + 1, ws.max_row + 1):
        category = str(ws.cell(row=r, column=col_category).value or "").strip()
        activity = str(ws.cell(row=r, column=col_activity).value or "").strip()
        sub_act  = str(ws.cell(row=r, column=col_subact).value or "").strip()

        if not category and not activity and not sub_act:
            continue  # blank row — skip silently, not an error
        if not category or not activity or not sub_act:
            skipped += 1
            errors.append(f"Row {r}: missing Category/Activity/Sub-Activity — skipped")
            continue

        planned_start = _cell_date_str(ws.cell(row=r, column=col_start).value) if col_start else None
        planned_end   = _cell_date_str(ws.cell(row=r, column=col_end).value) if col_end else None
        assigned_to   = str(ws.cell(row=r, column=col_assigned).value or "").strip() or None if col_assigned else None

        key = (category, activity, sub_act)
        bucket = existing_by_key.get(key)

        if bucket:
            existing_task = bucket.pop(0)   # consume one match — repeats in the file become new rows
            changing = (existing_task.planned_start != planned_start or existing_task.planned_end != planned_end
                        or existing_task.assigned_to != assigned_to)
            if changing:
                _new_task_version(db, existing_task, planned_start=planned_start, planned_end=planned_end, assigned_to=assigned_to)
                updated += 1
        else:
            t = GanttTask(
                project_id=project_id, category=category, activity=activity, task_name=sub_act,
                assigned_to=assigned_to, planned_start=planned_start, planned_end=planned_end,
                status="Not Started", percent_complete=0, sort_order=max_sort,
                version=1, is_current=True,
            )
            db.add(t)
            db.flush()
            t.entity_id = t.id
            max_sort += 1
            created += 1

    project.wbs_uploaded = True
    db.commit()

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],   # cap to avoid a huge payload on very messy files
        "total_errors": len(errors),
        "sheet_used": ws.title,
        "header_row": header_row,
    }


@gantt_router.get("/{project_id}/access-debug")
def access_debug(
    project_id: int, resource_id: int, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    """
    Management-only diagnostic: shows exactly which tier (draft / actual /
    plan / Allocation) is authoritative for this project this month, and what
    role each tier sees for the given resource — for troubleshooting why
    someone has view but not edit access, or no access at all.
    """
    if not is_management(current_resource):
        raise HTTPException(status_code=403, detail="Management only")

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    target = db.query(Resource).filter(Resource.id == resource_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Resource not found")

    from app.auth import _draft_staff_map, _current_month_staff_maps, _allocation_role, PROJECT_MANAGER_ROLES

    draft_map = _draft_staff_map(db)
    actual_map, plan_map = _current_month_staff_maps()
    code = project.project_code

    draft_role  = draft_map.get(code, {}).get(target.name) if code else None
    actual_role = actual_map.get(code, {}).get(target.name) if (actual_map is not None and code) else None
    plan_role   = plan_map.get(code, {}).get(target.name) if (plan_map is not None and code) else None
    alloc_role  = _allocation_role(target, project)

    if code in draft_map:
        authoritative, effective_role = "draft", draft_role
    elif actual_map is not None and code in actual_map:
        authoritative, effective_role = "actual_resource", actual_role
    elif plan_map is not None and code in plan_map:
        authoritative, effective_role = "plan_resource", plan_role
    else:
        authoritative, effective_role = "allocation (PMO unreachable or no PMO data at all)", alloc_role

    return {
        "project_code": code,
        "resource_name": target.name,
        "authoritative_tier": authoritative,
        "effective_role": effective_role,
        "can_edit": (effective_role or "").strip().lower() in {r.lower() for r in PROJECT_MANAGER_ROLES},
        "roles_by_tier": {
            "draft":            draft_role,
            "actual_resource":  actual_role,
            "plan_resource":    plan_role,
            "allocation_table": alloc_role,
        },
        "pmo_reachable": actual_map is not None,
    }


@gantt_router.get("/{project_id}/my-access")
def my_access(
    project_id: int, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    """
    Single source of truth for 'can the CURRENT user edit/view this
    project's Schedule' — the frontend calls this directly instead of
    keeping its own copy of the permission logic, so the two can never
    silently drift apart the way they just did.
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return {
        "can_view": can_view_schedule(current_resource, project, db),
        "can_edit": can_edit_schedule(current_resource, project, db),
    }

# ── Milestones — now live entirely in the Schedule section ────────────────────
# Moved out of Conversion/Plan/Actual: no more monthly push cycle, no more
# amount auto-calculated from percentage. Both are keyed in manually and are
# editable at any time, gated by the same Schedule permission as WBS tasks
# (Management + the project's Project Manager can edit; anyone else with
# Schedule access can view only).

class MilestoneIn(BaseModel):
    label:      str
    percentage: float = 0
    amount:     float = 0
    due_date:   Optional[str] = None
    is_completed: bool = False


class MilestoneUpdate(BaseModel):
    label:          Optional[str] = None
    percentage:     Optional[float] = None
    amount:         Optional[float] = None
    due_date:       Optional[str] = None
    is_completed:   Optional[bool] = None
    completed_date: Optional[str] = None
    invoice_number: Optional[str] = None
    invoice_date:   Optional[str] = None


def milestone_to_dict(m: Milestone) -> dict:
    return {
        "id": m.entity_id,   # stable across versions — never the per-version surrogate m.id
        "project_id": m.project_id, "label": m.label,
        "percentage": m.percentage, "amount": m.amount, "due_date": m.due_date,
        "is_completed": m.is_completed, "completed_date": m.completed_date,
        "invoice_number": m.invoice_number, "invoice_date": m.invoice_date,
        "version": m.version,
        "effective_date": m.effective_date.isoformat() if m.effective_date else None,
    }


def _get_current_milestone(db: Session, entity_id: int) -> Optional[Milestone]:
    return db.query(Milestone).filter(Milestone.entity_id == entity_id, Milestone.is_current == True).first()


def _new_milestone_version(db: Session, current: Milestone, **changes) -> Milestone:
    """SCD Type 2 write — same pattern as _new_task_version, see there for the full explanation."""
    now = datetime.utcnow()
    current.is_current = False
    current.end_date = now

    new = Milestone(
        entity_id=current.entity_id,
        version=current.version + 1,
        effective_date=now,
        end_date=None,
        is_current=True,
        project_id=current.project_id,
        label=changes.get("label", current.label),
        percentage=changes.get("percentage", current.percentage),
        amount=changes.get("amount", current.amount),
        due_date=changes.get("due_date", current.due_date),
        is_completed=changes.get("is_completed", current.is_completed),
        completed_date=changes.get("completed_date", current.completed_date),
        invoice_number=changes.get("invoice_number", current.invoice_number),
        invoice_date=changes.get("invoice_date", current.invoice_date),
        autocount_ref=current.autocount_ref,
        is_paid=current.is_paid,
        paid_date=current.paid_date,
        notification_sent=current.notification_sent,
    )
    db.add(new)
    return new


@gantt_router.get("/{project_id}/milestones")
def list_milestones(
    project_id: int, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    _project_for_view_or_403(project_id, db, current_resource)
    rows = (
        db.query(Milestone)
        .filter(Milestone.project_id == project_id, Milestone.is_current == True)
        .order_by(Milestone.entity_id)
        .all()
    )
    return {
        "milestones": [milestone_to_dict(m) for m in rows],
        "total_percentage": round(sum(m.percentage or 0 for m in rows), 2),
    }


@gantt_router.post("/{project_id}/milestones", status_code=201)
def create_milestone(
    project_id: int, body: MilestoneIn, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    require_schedule_edit_access(current_resource, project, db)

    # project.milestones is already filtered to is_current=True at the model
    # level, so this naturally only sums the CURRENT set — no change needed
    # here even though the underlying table now holds historical versions too.
    existing_total = sum(m.percentage or 0 for m in project.milestones)
    if existing_total + (body.percentage or 0) > 100.01:
        raise HTTPException(
            status_code=400,
            detail=f"Milestone percentages would total {existing_total + body.percentage:.1f}% — must not exceed 100%."
        )

    m = Milestone(
        project_id=project_id, label=body.label, percentage=body.percentage,
        amount=body.amount, due_date=body.due_date or None, is_completed=body.is_completed,
        version=1, is_current=True,
    )
    db.add(m)
    db.flush()      # assigns m.id so it can become its own stable entity_id
    m.entity_id = m.id
    db.commit()
    db.refresh(m)
    return milestone_to_dict(m)


@gantt_router.patch("/milestones/{milestone_id}")
def update_milestone(
    milestone_id: int, body: MilestoneUpdate, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    """milestone_id here is the stable entity_id, not a specific version's surrogate id."""
    m = _get_current_milestone(db, milestone_id)
    if not m:
        raise HTTPException(status_code=404, detail="Milestone not found")
    project = db.query(Project).filter(Project.id == m.project_id).first()
    if project:
        require_schedule_edit_access(current_resource, project, db)

    if body.percentage is not None and project:
        existing_total = sum(x.percentage or 0 for x in project.milestones if x.entity_id != milestone_id)
        if existing_total + body.percentage > 100.01:
            raise HTTPException(
                status_code=400,
                detail=f"Milestone percentages would total {existing_total + body.percentage:.1f}% — must not exceed 100%."
            )

    changes = {}
    for field in ["label", "percentage", "amount", "due_date", "is_completed",
                  "completed_date", "invoice_number", "invoice_date"]:
        val = getattr(body, field)
        if val is not None:
            changes[field] = val

    # Auto-fill completed_date the moment is_completed flips true, if not already set
    if body.is_completed and not m.completed_date and "completed_date" not in changes:
        changes["completed_date"] = date.today().isoformat()

    if changes:
        m = _new_milestone_version(db, m, **changes)

    db.commit()
    db.refresh(m)
    return milestone_to_dict(m)


@gantt_router.delete("/milestones/{milestone_id}", status_code=204)
def delete_milestone(
    milestone_id: int, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    """Soft delete — same reasoning as delete_task. milestone_id is the entity_id."""
    m = _get_current_milestone(db, milestone_id)
    if not m:
        raise HTTPException(status_code=404, detail="Milestone not found")
    project = db.query(Project).filter(Project.id == m.project_id).first()
    if project:
        require_schedule_edit_access(current_resource, project, db)
    m.is_current = False
    m.end_date = datetime.utcnow()
    db.commit()
    return None

# ── Pending date change approvals — Management only ────────────────────────────

@gantt_router.get("/{project_id}/pending-date-changes")
def list_pending_date_changes(
    project_id: int, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    if not is_management(current_resource):
        raise HTTPException(status_code=403, detail="Management only")
    rows = (
        db.query(PendingDateChange)
        .filter(PendingDateChange.project_id == project_id, PendingDateChange.status == "pending")
        .order_by(PendingDateChange.requested_at)
        .all()
    )
    out = []
    for p in rows:
        d = pending_date_change_to_dict(p)
        task = _get_current_task(db, p.gantt_task_id)
        d["task_name"] = task.task_name if task else None
        d["category"] = task.category if task else None
        d["activity"] = task.activity if task else None
        out.append(d)
    return {"pending": out}


class DateChangeReview(BaseModel):
    note: Optional[str] = None


@gantt_router.post("/pending-date-changes/{pending_id}/approve")
def approve_date_change(
    pending_id: int, body: DateChangeReview, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    if not is_management(current_resource):
        raise HTTPException(status_code=403, detail="Management only")
    p = db.query(PendingDateChange).filter(PendingDateChange.id == pending_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Request not found")
    if p.status != "pending":
        raise HTTPException(status_code=400, detail=f"This request is already {p.status}")

    t = _get_current_task(db, p.gantt_task_id)
    if not t:
        raise HTTPException(status_code=404, detail="Task no longer exists")

    t = _new_task_version(db, t, planned_start=p.requested_planned_start, planned_end=p.requested_planned_end)

    p.status = "approved"
    p.reviewed_by = current_resource.id if current_resource else None
    p.reviewed_at = datetime.utcnow()
    p.review_note = body.note

    db.commit()
    db.refresh(t)
    return task_to_dict(t)


@gantt_router.post("/pending-date-changes/{pending_id}/reject")
def reject_date_change(
    pending_id: int, body: DateChangeReview, db: Session = Depends(get_db),
    current_resource: Optional[Resource] = Depends(get_current_resource),
):
    if not is_management(current_resource):
        raise HTTPException(status_code=403, detail="Management only")
    p = db.query(PendingDateChange).filter(PendingDateChange.id == pending_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Request not found")
    if p.status != "pending":
        raise HTTPException(status_code=400, detail=f"This request is already {p.status}")

    p.status = "rejected"
    p.reviewed_by = current_resource.id if current_resource else None
    p.reviewed_at = datetime.utcnow()
    p.review_note = body.note

    db.commit()
    return {"message": "Request rejected"}
